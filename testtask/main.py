import mimetypes
import os
import openpyxl

from openpyxl import load_workbook
from typing import List, Tuple


class Config:
    def __init__(self, 
                symlink: bool = False, 
                excel_name: str = 'result.xlsx', 
                is_exist: bool = False, 
                root_path: str = '.', 
                file_extensions: List[str] = ['txt'],
                mime_types: List[str] = ['text/xml', 'text/csv', 'text/plain'],
                all_extensions: bool = True,
                all_mimes: bool = True,
        ) -> None:
        self.symlink = symlink
        self.excel_path = os.getcwd() + "/" + excel_name
        self.excel_name = excel_name
        self.is_exist = is_exist
        self.root_path = root_path
        self.file_extensions = file_extensions
        self.mime_types = mime_types
        self.all_extensions = all_extensions
        self.all_mimes = all_mimes

class ExcelController:
    def __init__(self, config: Config = Config()) -> None:
        self.config = config
        self.workbook = openpyxl.Workbook()
        self.name = config.excel_name
        self.init()

    def init(self):
        if not os.path.exists(self.config.excel_path):
            self.workbook.save(self.name)
        self.workbook = load_workbook(self.config.excel_path)

    def add(self, file_name: str, path: str = None):
        ws1 = self.workbook.active

        rows_length = len(ws1['A'])
        extension = file_name.split('.')
        if len(extension) == 1:
            extension = None
        else:
            extension = extension[-1]

        ws1.append([rows_length, path, file_name, extension])
        self.workbook.save(self.config.excel_path)
    
    def is_exist():
        #TODO
        pass


class Traverser:
    def __init__(self, config = Config()) -> None:
        self.config = config

    def traverse(self) -> Tuple[str, str]:
        for root, _, files in os.walk(self.config.root_path, topdown=True):
            for name in files:
                path = os.path.join(root, name)

                if self.check_file(path):
                    yield (root, name)

    def check_file(self, path: str):
        name = path.split('/')[-1]

        if name.startswith('.'):
            return False

        if self.config.all_extensions or name.split('.')[-1] in self.config.file_extensions:
            if self.config.all_mimes or mimetypes.guess_type(path) in self.config.mime_types:
                return True

        return False


if __name__ == "__main__":
    tr = Traverser()
    excel = ExcelController()

    for dir, name in tr.traverse():
        excel.add(name, dir)

